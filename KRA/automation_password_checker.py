import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from playwright.sync_api import sync_playwright
from PIL import Image
from io import BytesIO
import tkinter as tk
from tkinter import PhotoImage, Label

# Step 1: Try to open an existing Excel file or create a new one
try:
    wb = load_workbook(r"C:\Users\DELL\Desktop\BCL AUTOMATIONS/passwords_clients.xlsx")
    ws = wb.active
except FileNotFoundError:
    print("The 'passwords_clients.xlsx' file does not exist.")
    exit(1)

# Define fill patterns for green and red
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

# Step 2: Define a function to check the validity of a password
def is_valid_password(password):
    # Check if the password is equal to "bcl123"
    if password == "bclitax2023":
        return True
    else:
        return False

# Step 3: Initialize Playwright
with sync_playwright() as p:
    browser = p.chromium.launch()
    page = browser.new_page()

    # Step 4: Iterate through the rows of the Excel file and interact with the website
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        client = row[0].value
        password = row[1].value

        if password is None:
            print(f"Skipping client '{client}' due to a missing password.")
            continue

        # Create a new Tkinter window for each iteration
        root = tk.Tk()
        root.title("CAPTCHA Image")

        # Create a Tkinter Label for displaying the image
        image_label = Label(root)
        image_label.pack()

        # Placeholder for PhotoImage object
        captcha_photo = None

        # Step 5: Use Playwright to open the website and fill in fields
        page.goto("https://itax.kra.go.ke/KRA-Portal/")
        page.fill("#logid", client)
        page.get_by_role("link", name="Continue").click()

        # Example: Fill in username and password fields
        page.fill("#userName", client)
        page.fill("#xxZTT9p2wQ", password)

        # Capture the CAPTCHA image
        captcha_img = page.locator("#captcha_img")
        screenshot = captcha_img.screenshot()

        # Save the screenshot to a file (use a unique file name for each iteration)
        screenshot_path = f"captcha_{row[0].row}.png"
        with open(screenshot_path, "wb") as screenshot_file:
            screenshot_file.write(screenshot)

        # Create a Tkinter PhotoImage from the saved file
        captcha_photo = PhotoImage(file=screenshot_path)

        # Update the Tkinter Label to display the new image
        image_label.config(image=captcha_photo)

        # Fill in security stamp
        sec_Stamp = input("Key in Security stamp Security Stamp: ")
        page.fill("#captcahText", sec_Stamp)

        # Click the login button to submit the form
        page.get_by_role("link", name="Login").click()

        # Check if the login is successful by looking for "My Profile" text
        if page.get_by_text("My Profile"):
            # Click the logout button
            page.get_by_role("link", name="Logout").click()
            
            # Update the status as valid
            status = "Valid"
        else:
            if page.get_by_role("row", name="Invalid Login Id or Password.", exact=True).get_by_role("cell") or page.get_by_role("cell", name="The account has been locked.", exact=True):
                
                # Update the status as invalid
                status = "Invalid"
        

        # Step 7: Update the "Status" field in the Excel file
        ws.cell(row=row[0].row, column=3, value=status)  # Update the "Status" cell directly

        # Step 8: Apply fill color based on status
        if status == "Valid":
            ws.cell(row=row[0].row, column=3).fill = green_fill
        else:
            ws.cell(row=row[0].row, column=3).fill = red_fill

        # Save the updated workbook to the Excel file
        wb.save("passwords_clients.xlsx")

        # Close the Tkinter window for this iteration
        root.destroy()

    # Close the browser
    browser.close()
    

#  Delete all images in the current directory  
for file in os.listdir():
    if file.endswith(".png") and file.startswith("captcha"):
        os.remove(file)
