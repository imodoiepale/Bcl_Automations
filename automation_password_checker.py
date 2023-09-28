import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from playwright.sync_api import sync_playwright

# Step 1: Try to open an existing Excel file or create a new one
try:
    wb = load_workbook("passwords_clients.xlsx")
    ws = wb.active
except FileNotFoundError:
    print("The 'passwords_clients.xlsx' file does not exist.")
    exit(1)

# Define fill patterns for green and red
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

# Step 2: Define a function to check the validity of a password
def is_valid_password(password):
    # Check if the password is equal to "bcl123"
    if password == "bclitax2023":
        return True
    else:
        return False

# Step 3: Initialize Playwright
with sync_playwright() as p:
    browser = p.chromium.launch()
    page = browser.new_page()

    # Step 4: Iterate through the rows of the Excel file and interact with the website
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        client = row[0].value
        password = row[1].value

        if password is None:
            print(f"Skipping client '{client}' due to missing password.")
            continue

        # Step 5: Use Playwright to open the website and fill in fields
        page.goto("https://itax.kra.go.ke/KRA-Portal/")  # Replace with the URL of the website you want to automate
        page.fill("#logid", client) #fill in username
        page.get_by_role("link", name="Continue").click()  #continue to next page
        
        # Example: Fill in username and password fields
        
        page.fill("#userName", client)
        page.fill("#xxZTT9p2wQ", password)
        
        # fill in security stamp
        sec_Stamp = input("Key in Security stamp Security Stamp : ")
        page.fill("#captcahText", sec_Stamp)
        
        # Click the login button to submit the form
        page.get_by_role("link", name="Login").click()

        ######################################soon work on previewing the security stamp on the page###################################
        
        page.click("#login-button")

        # Step 6: Capture the result on the website
        # You will need to inspect the website's HTML to locate the element containing the result

        # Example: Assuming the result is displayed in a div with id "result"
        result =  page.get_by_role("row", name="Invalid Login Id or Password.", exact=True).locator("b")

        # Step 7: Update the "Status" field in the Excel file
        status = "Not Valid" if is_valid_password(result) else "Valid"
        ws.cell(row=row[0].row, column=3, value=status)  # Update the "Status" cell directly

        # Step 8: Apply green fill for valid and red fill for not valid
        if status == "Valid":
            ws.cell(row=row[0].row, column=3).fill = green_fill
        else:
            ws.cell(row=row[0].row, column=3).fill = red_fill

    # Step 9: Save the updated workbook to the Excel file
    wb.save("passwords_clients.xlsx")

    # Step 10: Close the browser
    browser.close()
