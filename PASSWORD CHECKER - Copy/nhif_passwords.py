import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from playwright.sync_api import sync_playwright
import time

# Get the current directory
current_directory = os.getcwd()

# Specify the file path relative to the current directory
file_path = os.path.join(current_directory, "nhif_passwords.xlsx")

try:
    wb = load_workbook(file_path)
    ws = wb.active
except FileNotFoundError:
    print("The 'nhif_passwords.xlsx' file does not exist.")
    exit(1)

# Define fill patterns for green and red
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

def login_and_update_status(username, password, row):
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()

        # Iterate through the rows of the Excel file and interact with the website
        for excel_row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=4):
            current_username = excel_row[0].value  # Index 0 corresponds to the second column (Log In User)
            current_password = excel_row[1].value  # Index 1 corresponds to the third column (Password)
            current_status = excel_row[2].value     # Index 2 corresponds to the fourth column (Status)

            # Print information for debugging
            print(f"Processing: {current_username}, {current_password}, {current_status}")

            time.sleep(1)
            # Navigate to the eCitizen login page
            page.goto("https://payrollbyproduct.nhif.or.ke/byproduct/login.php")
            
            # Fill in the login form
            page.fill("#logUsername", str(current_username))
            page.fill("#logPassword", str(current_password))

            # Click the login button to submit the form
            page.get_by_role("button", name="Login").click()

            try:
                # Check if the login is successful by looking for "Byproduct Upload" text
                page.get_by_text("Byproduct Upload")
                # Click the logout button
                page.get_by_role("link", name="Logout").click()
                # Update the status as valid
                current_status = "Valid"

            except:
                time.sleep(1)
                # Handle error if login fails
                print("Login failed.")
                # Update the status as invalid
                current_status = "Invalid"
                pass

            # Print information for debugging
            print(f"Updated Status: {current_status}")

            # Update Excel file with status and color
            ws.cell(row=excel_row[0].row, column=4, value=current_status)  # Update the "Status" cell directly

            # Apply fill color based on status
            if current_status == "Valid":
                ws.cell(row=excel_row[0].row, column=4).fill = green_fill
            else:
                ws.cell(row=excel_row[0].row, column=4).fill = red_fill

            # Save the updated workbook to the Excel file
            wb.save("nhif_passwords.xlsx")

        print("Iteration complete")
        # Close the browser
        browser.close()

# Example usage with a specific row
login_and_update_status("sample_username", "sample_password", 2)